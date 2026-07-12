# -*- coding: utf-8 -*-
"""
Extraction d'un classeur Excel du professeur : formules + valeurs cachées.

C'est l'outil qui a servi a documenter la feuille « gramme » (Issues.md #4) et
a construire les oracles (excel_twin*.py). Il est commite pour que N'IMPORTE
QUEL mainteneur futur puisse depouiller un nouveau classeur sans reinventer la
procedure — voir docs/MAINTENANCE.md, section « Nouveau classeur Excel ».

Usage (openpyxl n'est PAS dans requirements — l'installer localement) :

    cd backend
    .venv\\Scripts\\pip install openpyxl        # une fois, local seulement
    .venv\\Scripts\\python ..\\tools\\extract_workbook.py "..\\Data\\MonClasseur.xlsx"
    .venv\\Scripts\\python ..\\tools\\extract_workbook.py "..\\Data\\MonClasseur.xlsx" --sheet "Feuil1" --out dump.txt

Sortie, pour chaque cellule non vide :
    [F] D43: =D22*(D41+D42)  => 174.24009260940232     (formule + valeur cachee)
    [N] D21: 0.73  => 0.73                             (nombre saisi)
    [T] B22: Proportion massique de liant (%)          (texte)

Les valeurs cachees (=> ...) sont celles que le classeur a memorisees au
dernier calcul Excel : ce sont ELLES qu'on epingle dans le CACHED d'un twin
(tolerance 1e-9), jamais des valeurs recalculees a la main.
"""

from __future__ import annotations

import argparse
import sys


def extraire(chemin: str, feuille: str | None = None, sortie=None) -> None:
    try:
        import openpyxl  # import local volontaire (hors requirements)
    except ImportError:
        print("openpyxl manquant. Installer localement :", file=sys.stderr)
        print("  backend\\.venv\\Scripts\\pip install openpyxl", file=sys.stderr)
        raise SystemExit(1)

    out = sortie or sys.stdout
    wb_f = openpyxl.load_workbook(chemin, data_only=False)  # formules
    wb_v = openpyxl.load_workbook(chemin, data_only=True)   # valeurs cachees

    print("=== FEUILLES ===", file=out)
    for ws in wb_f.worksheets:
        print(f"  {ws.title!r}  dims={ws.dimensions}  max_row={ws.max_row} "
              f"max_col={ws.max_column}", file=out)

    feuilles = [f for f in wb_f.worksheets if feuille is None or f.title == feuille]
    if feuille is not None and not feuilles:
        print(f"Feuille introuvable : {feuille!r}", file=sys.stderr)
        raise SystemExit(2)

    for ws_f in feuilles:
        ws_v = wb_v[ws_f.title]
        print(f"\n===== FEUILLE {ws_f.title!r} =====", file=out)
        for row in ws_f.iter_rows():
            for cell in row:
                f = cell.value
                if f is None:
                    continue
                v = ws_v[cell.coordinate].value
                est_formule = isinstance(f, str) and f.startswith("=")
                tag = "F" if est_formule else ("T" if isinstance(f, str) else "N")
                fs = str(f).replace("\n", " ")
                vs = "" if v is None else f"  => {v!r}"
                print(f"  [{tag}] {cell.coordinate}: {fs}{vs}", file=out)


def main() -> None:
    p = argparse.ArgumentParser(description="Dump formules + valeurs cachees d'un classeur.")
    p.add_argument("classeur", help="Chemin du .xlsx")
    p.add_argument("--sheet", default=None, help="Nom d'une feuille (defaut : toutes)")
    p.add_argument("--out", default=None, help="Fichier de sortie (defaut : stdout)")
    args = p.parse_args()

    if args.out:
        with open(args.out, "w", encoding="utf-8") as fh:
            extraire(args.classeur, args.sheet, fh)
        print(f"Dump ecrit : {args.out}")
    else:
        extraire(args.classeur, args.sheet)


if __name__ == "__main__":
    main()
